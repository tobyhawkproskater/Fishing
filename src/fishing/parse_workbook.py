"""Parse Salmon Steelhead Trout.xlsx into structured dicts."""
from __future__ import annotations

from datetime import date, datetime, timedelta

from openpyxl import load_workbook

from . import SOURCES

# Excel serial date origin (Windows). Excel incorrectly treats 1900 as a leap year,
# so for serials >= 60 we subtract one day. Origin is 1899-12-30 to compensate.
_EXCEL_EPOCH = datetime(1899, 12, 30)


def _excel_serial_to_date(serial: float) -> date | None:
    try:
        return (_EXCEL_EPOCH + timedelta(days=float(serial))).date()
    except (ValueError, TypeError, OverflowError):
        return None


def _rows(ws, header_row: int):
    headers = [c.value for c in ws[header_row]]
    out = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v in (None, "") for v in row):
            continue
        out.append({h: v for h, v in zip(headers, row) if h is not None})
    return [h for h in headers if h is not None], out


def _opens_with(d: dict, prefix: str) -> bool:
    first_key = next(iter(d), "")
    val = d.get(first_key)
    return isinstance(val, str) and val.strip().lower().startswith(prefix)


def parse() -> dict:
    wb = load_workbook(str(SOURCES["workbook"]), data_only=True)

    species_headers, species = _rows(wb["Local Fishing Info"], 1)
    calendar_headers, calendar = _rows(wb["Calendar"], 2)
    gear_headers, gear = _rows(wb["Gear Tracker"], 2)
    state_headers, state_master = _rows(wb["State Master"], 1)

    # Salmon Log 2025: header in row 2, totals row at the bottom — split them out.
    log_headers, log_rows_raw = _rows(wb["Salmon Log 2025"], 2)
    log = []
    totals = None
    for r in log_rows_raw:
        first_val = r.get("Date")
        if isinstance(first_val, str) and first_val.strip().lower() in {"totals", "hours per fish", "hours per week"}:
            if first_val.strip().lower() == "totals":
                totals = r
            continue
        if isinstance(first_val, (int, float)):
            d = _excel_serial_to_date(first_val)
            if d:
                r["Date"] = d.isoformat()
        if r.get("Fish") is None and r.get("Hours") is None:
            continue
        log.append(r)

    # PROBE sheet (depth/line tables) — keep raw rows for now; the structure is
    # specific and we want the future MCP server to expose a small lookup helper.
    probe_headers, probe_rows = _rows(wb["PROBE"], 2)

    return {
        "species": {"headers": species_headers, "rows": species},
        "calendar": {"headers": calendar_headers, "rows": calendar},
        "gear": {"headers": gear_headers, "rows": gear},
        "log_2025": {"headers": log_headers, "rows": log, "totals": totals},
        "state_master": {"headers": state_headers, "rows": state_master},
        "probe": {"headers": probe_headers, "rows": probe_rows},
    }


if __name__ == "__main__":
    import json

    data = parse()
    print(json.dumps({k: {"count": len(v["rows"])} for k, v in data.items()}, indent=2))
