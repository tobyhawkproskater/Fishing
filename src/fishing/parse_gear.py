"""Parse the standalone `Fishing Gear.xlsx` catalog."""
from __future__ import annotations

from openpyxl import load_workbook

from . import SOURCES


def _rows(ws, header_row: int):
    headers = [c.value for c in ws[header_row]]
    out = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v in (None, "") for v in row):
            continue
        out.append({h: v for h, v in zip(headers, row) if h is not None})
    return [h for h in headers if h is not None], out


def parse_gear() -> dict:
    """Parse `Fishing Gear.xlsx`.

    Schema (header row 2): Use | Purpose | Brand | Model | Type | Number |
    Cost (rod and reel) | Length | Power | Taper | Line Rating | Lure Rating |
    Troll Rating | Reel | Line | Guests | Notes | Location.

    A trailing totals row (only the Cost column populated) is split out as
    `totals` so the per-rod rows stay clean.
    """
    wb = load_workbook(str(SOURCES["gear_workbook"]), data_only=True)
    headers, rows = _rows(wb["Gear Tracker"], 2)
    catalog: list[dict] = []
    totals: dict | None = None
    for r in rows:
        rod_identifying = any(
            r.get(k) for k in ("Use", "Brand", "Model", "Number")
        )
        if not rod_identifying:
            if r.get("Cost (rod and reel)") is not None:
                totals = r
            continue
        catalog.append(r)
    return {"headers": headers, "rows": catalog, "totals": totals}


if __name__ == "__main__":
    import json
    data = parse_gear()
    print(json.dumps({"rods": len(data["rows"]), "totals": data["totals"]}, indent=2, default=str))
